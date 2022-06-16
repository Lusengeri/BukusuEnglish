from openpyxl import load_workbook 

def splitter(input_string):
    entries = []
    test_string = input_string
    space_count = test_string.count("   ")

    if space_count == 0:
        entries.append(test_string)
    else:
        while space_count >= 2:
            if space_count == 2:
                entries.append(test_string)
                break
            else:
                first_breakdown = test_string.split("   ")
                second_breakdown = first_breakdown[2].split(" ")
                second_breakdown_len = len(second_breakdown)
                if (second_breakdown[second_breakdown_len-1]) == '(?)':
                    second_term = second_breakdown[second_breakdown_len-2] + " (?)"
                else:
                    second_term = second_breakdown[second_breakdown_len-1]

                index_of_second_term = test_string.find(second_term)
                entries.append(test_string[0:index_of_second_term])
                test_string = test_string[index_of_second_term:]
                space_count = test_string.count("   ")

    return entries 

def joiner(strings):
    out_strings = [] 
    for string in strings:
        current_string_index = strings.index(string)
        space_count = string.count("   ")
        if space_count == 0:
            if current_string_index == 0:
                continue
            else:
                curr_len = len(out_strings)
                out_strings[curr_len-1] = out_strings[curr_len-1] + (" " + string)
        else:
            out_strings.append(string)

    return out_strings


output_file = open('all_entries.sql', 'w+')
output_file.write('CREATE TABLE definitions (_id INTEGER PRIMARY KEY, word text, pos text, definition text, unaccented text);\n')

workbook = load_workbook(filename='input_dictionary.xlsx')
ws = workbook['Table 1']

cell_range = ws['A1': 'A4643']
for row in cell_range:
    curr_index = cell_range.index(row)
    curr_value = row[0].value 
    if curr_value == "" or None:
        continue

    if curr_index < 4605:
        next_line = cell_range[curr_index+1][0].value
        if next_line.count("   ") == 0:
            row[0].value += " " + next_line
            cell_range[curr_index+1][0].value = ""

    #Split row[0] and write to file
    print(curr_index+1)
    entries = splitter(row[0].value)
    for entry in entries:
        word, pos, definition = entry.split("   ")
        word.strip()
        pos.strip()
        definition.strip()
        unaccented = word.replace('á', "a").replace("ú", "u").replace("é", "e").replace("í", "i").replace(".", "").replace("ii", "i").replace("aa", "a").replace("uu", "u").replace("ó","o").replace("(?)", "")
        query = "INSERT INTO definitions (word, pos, definition, unaccented) VALUES ('{0}', '{1}', '{2}', '{3}');\n".format(word, pos, definition, unaccented)
        output_file.write(query)

output_file.close()
